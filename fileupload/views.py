import os
from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.core.files.storage import default_storage
import openpyxl
import requests
import logging
from bs4 import BeautifulSoup
from openpyxl import Workbook
from django.http import JsonResponse
import time
import shutil
from django.http import FileResponse
# Initialize logger
logger = logging.getLogger(__name__)

# View to serve the HTML page
def serve_html(request):
    file_path = os.path.join(os.path.dirname(__file__), "index.html")
    with open(file_path, "r") as file:
        html_content = file.read()
    return HttpResponse(html_content)


@api_view(["POST"])
def upload_file(request):
    if "file" not in request.FILES:
        return Response(
            {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
        )

    file = request.FILES["file"]

    # Validate the file type (only .xlsx allowed)
    if not file.name.endswith(".xlsx"):
        return Response(
            {"error": "Invalid file type. Only .xlsx files are allowed."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Save the uploaded file to the 'upload' folder
    upload_folder = os.path.join(settings.MEDIA_ROOT, "uploads")
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)

    file_name = os.path.join(upload_folder, file.name)
    with open(file_name, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)

    # Process the Excel file synchronously
    try:
        asin_list = process_excel_file(file_name)
        processed_file = scrape_asins(asin_list)  # Pass the asin_list directly
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response(
        {
            "success": True,
            "message": "File uploaded and processed successfully",
            "processed_file": processed_file["file_path"],
        }
    )


def process_excel_file(file_path):
    asin_list = []
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Assuming ASINs are in the second column (column index 1)
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            asin = row[0]  # Get the ASIN from the second column
            if asin:
                asin_list.append(asin)

        wb.close()
    except Exception as e:
        raise Exception(f"Error processing Excel file: {e}")

    return asin_list


def safe_get_text(element, default_value="N/A"):
    """Safely gets text from an element if it exists, otherwise returns a default value."""
    return element.get_text(strip=True) if element else default_value


def scrape_asins(asin_list):
    excel_data = []

    for i, asin in enumerate(asin_list):
        if not asin:
            logger.error(f"ASIN is missing for item {i}. Skipping...")
            continue

        try:
            logger.info(f"Processing ASIN: {asin}")

            url = "https://www.amazon.de/acp/buffet-mobile-card/buffet-mobile-card-3e67eb5a-92a5-4eae-9a4d-c1d3082690fb-1734571386882/getRspManufacturerContent?page-type=DetailAW&stamp=1734623286402"

            headers = {
                "accept": "text/html, application/json",
                "accept-language": "en-GB,en;q=0.9,be;q=0.8,ur;q=0.7",
                "content-type": "application/json",
                "device-memory": "8",
                "downlink": "4.25",
                "dpr": "2",
                "ect": "4g",
                "priority": "u=1, i",
                "rtt": "250",
                "sec-ch-device-memory": "8",
                "sec-ch-dpr": "2",
                "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
                "sec-ch-ua-mobile": "?1",
                "sec-ch-ua-platform": '"Android"',
                "sec-ch-ua-platform-version": '"6.0"',
                "sec-ch-viewport-width": "1145",
                "x-amz-acp-params": "tok=FBsk2BFo33RUH3sujiaU_dkdakUcEBnthvUxK3jaTj4;ts=1734623286395;rid=YPAQAPMK7HS057YPN4AD;d1=711;d2=0",
                "x-amz-amabot-click-attributes": "disable",
                "x-requested-with": "XMLHttpRequest",
                "cookie": 'session-id=261-5758951-0539711; session-id-time=2082787201l; i18n-prefs=EUR; lc-acbde=en_GB; sp-cdn="L5Z9:PK"; ubid-acbde=261-5393323-8128104; session-token=RVuGuCOz7rQrxfHb0cosNpD+u0bC7roD/2RaAnDtCXh9SGiSIzUEOGPNsdMo2/H607FyEYsyMy+zh8u/i3tXuhqUwki7bkMx1KYf8OFrr2SJsalca8qxe10aZmm1dq7UEZS1hA2CdN9EWE2sQGmHnBWb84YWuoPtFhBCv5BZGpWM42S8PYSiGlorZaav0JYEgUqVWCpJZpB13sq6Guy8C9wIrEjHGn2EtYaCj8PQiyZpQTF7qHQub3QSq517SaSOk+j8adBQPOeCOakcSgveJjTU/9y6sOi00KHadgZG4/x7rs5jm+ItnQBK1JoS81IGX2nsX4gCLycCjInxx9FUXE17K9oU4wil',
                "Referer": "https://www.amazon.de/dp/B0BJ1Q3HWZ?th=1",
                "Referrer-Policy": "strict-origin-when-cross-origin",
            }

            request_body = {"asin": asin}
            response = requests.post(url, json=request_body, headers=headers)
            soup = BeautifulSoup(response.text, "html.parser")

            # Use the safe_get_text() function to handle missing elements gracefully
            manufacturer_info = {
                "name": safe_get_text(
                    soup.select_one(".a-box .a-box-inner .a-size-base.a-text-bold")
                ),
                "address": ", ".join(
                    [
                        safe_get_text(
                            soup.select(".a-box .a-box-inner .a-list-item")[i]
                        )
                        for i in range(1, 3)
                    ]
                    if len(soup.select(".a-box .a-box-inner .a-list-item")) > 1
                    else ["N/A"]
                ),
                "email": safe_get_text(
                    soup.select_one(
                        ".a-box .a-box-inner .a-spacing-top-small .a-list-item"
                    )
                ),
            }

            eu_responsible_person = {
                "name": safe_get_text(
                    soup.select_one(".a-box .a-box-inner .a-size-base.a-text-bold")
                ),
                "address": ", ".join(
                    [
                        safe_get_text(
                            soup.select(".a-box .a-box-inner .a-list-item")[i]
                        )
                        for i in range(1, 4)
                    ]
                    if len(soup.select(".a-box .a-box-inner .a-list-item")) > 1
                    else ["N/A"]
                ),
                "email": safe_get_text(
                    soup.select_one(
                        ".a-box .a-box-inner .a-spacing-top-small .a-list-item"
                    )
                ),
            }

            excel_data.append(
                {
                    "ASIN": asin,
                    "Manufacturer Information.name": manufacturer_info["name"],
                    "Manufacturer Information.address": manufacturer_info["address"],
                    "Manufacturer Information.email": manufacturer_info["email"],
                    "EU Responsible Person.name": eu_responsible_person["name"],
                    "EU Responsible Person.address": eu_responsible_person["address"],
                    "EU Responsible Person.email": eu_responsible_person["email"],
                }
            )

            logger.info(f"Data retrieved for ASIN: {asin}")

        except Exception as e:
            logger.error(f"Error processing ASIN: {asin} - {e}")

    # Save the processed file to 'files' folder
    timestamp = str(int(time.time()))
    file_name = f"{timestamp}.xlsx"
    output_folder = os.path.join(settings.MEDIA_ROOT, "files")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    output_path = os.path.join(output_folder, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "ASIN Data"

    # Write headers from the excel_data list
    headers = list(excel_data[0].keys()) if excel_data else []
    ws.append(headers)

    for data in excel_data:
        # Ensure data is properly encoded (UTF-8) before appending to Excel
        row = [str(value).encode('utf-8').decode('utf-8') for value in data.values()]
        ws.append(row)

    wb.save(output_path)

    # Return the file path for downloading
    file_path = file_name
    logger.info(f"Excel file created at {file_path}")

    return {"file_path": file_path}




def download_file(request, filename):
    # Fix the file path by removing the leading slash
    file_path = os.path.join(settings.MEDIA_ROOT, 'files', filename)

    # Check if the file exists
    if os.path.exists(file_path):
        # Use FileResponse, Django will handle opening and closing the file
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
    
    return JsonResponse({'success': False, 'message': 'File not found.'})
